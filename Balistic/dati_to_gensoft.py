import pandas as pd
import math
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def text_before(text, delimiter="-"):
    """Replicates Excel's TEXTBEFORE function."""
    if pd.isna(text):
        return ""
    parts = str(text).split(delimiter)
    return parts[0] if parts else ""

def text_after(text, delimiter="-"):
    """Replicates Excel's TEXTAFTER function."""
    if pd.isna(text):
        return ""
    parts = str(text).split(delimiter)
    return parts[1] if len(parts) > 1 else ""

def map_gender_bg(gender):
    """Maps English gender to Bulgarian gender prefix."""
    gender_map = {
        "MENS": "Мъжки",
        "WOMENS": "Дамски",
        "ADULT UNISEX": "Унисекс",
        "BOYS": "Детски",
        "GIRLS": "Детски",
        "GRD SCHOOL UNS": "Детски",
        "GRD SCHOOL UNSX": "Унисекс",
        "BOYS GRADE SCHL": "Детски",
        "Youth unisex": "Детски"
    }
    return gender_map.get(gender, "Унисекс")  # Default to "Унисекс" if gender not found

def map_silhouette_bg(silhouette):
    """Maps English silhouette to Bulgarian product type."""
    silhouette_map = {
        "BACKPACK": "Раница",
        "SHORT SLEEVE T-SHIRT": "Тениска",
        "LOW TOP": "Маратонки",
        "CREW SOCK": "Чорапи",
        "NO SHOW SOCK": "Чорапи",
        "HOODED LONG SLEEVE TOP": "Суитшърт с качулка",
        "FULL LENGTH PANT": "Панталон",
        "ANKLE LENGTH PANT": "Панталон",
        "HOODED FULL ZIP LS TOP": "Суитшърт с качулка и цип"
    }
    return silhouette_map.get(silhouette, "")  # Default to empty string if not found

def transform_dati_imp_to_gensoft(input_file, output_file):
    """
    Reads an Excel file, transforms the data into Gensoft format, and writes
    a new Excel file with a two-row header matching the required format.
    """
    print(f"Reading input file: {input_file}")
    
    # 1) Read the first sheet
    try:
        xls = pd.ExcelFile(input_file)
        sheet_name = xls.sheet_names[0]  # Automatically selects the first sheet
        dati_imp = pd.read_excel(xls, sheet_name=sheet_name)
        print(f"Successfully read sheet '{sheet_name}' with {len(dati_imp)} rows.")
    except Exception as e:
        print(f"Error reading sheet: {str(e)}")
        raise
    
    # 2) Dynamic column detection/renaming
    expected_columns = {
        'Art.num': ['Art.num', 'Article Number', 'Product Code'],
        'SizeConverted': ['SizeConverted', 'Size', 'Converted Size'],
        'Description': ['Description', 'Product Description'],
        'Season': ['Season', 'Year/Season', 'Seasons'],
        'Barcode': ['Barcode', 'EAN Code'],
        'Box BarCode': ['Box BarCode', 'Box Barcode', 'Box Bar Code'],
        'Dlv.qty': ['Dlv.qty', 'Quantity Delivered'],
        'FPC Price w/o VAT in BGN': ['FPC Price w/o VAT in BGN', 'Net Price'],
        'Division': ['Division', 'Category'],
        'Batch': ['Batch', 'Batch Number'],
        'Material Content': ['Material Content', 'Composition'],
        'Gender': ['Gender', 'Sex'],
        'Silhouette': ['Silhouette', 'Product Type']
    }
    
    column_mapping = {}
    for key, possible_names in expected_columns.items():
        for name in possible_names:
            if name in dati_imp.columns:
                column_mapping[name] = key
                break
    
    dati_imp = dati_imp.rename(columns=column_mapping)
    missing_columns = [col for col in expected_columns.keys() if col not in dati_imp.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # 3) Create intermediate dataframe (dati) with transformations
    dati = pd.DataFrame()
    dati['CODICE NIKE+COLOR'] = dati_imp['Art.num']
    dati['COD NIKE'] = dati['CODICE NIKE+COLOR'].apply(text_before)
    dati['COD COLOR'] = dati['CODICE NIKE+COLOR'].apply(text_after)
    dati['TAGLIA'] = dati_imp['SizeConverted']
    dati['CODICE BALLISTIC'] = dati.apply(
        lambda row: f"{row['COD NIKE']}-{row['COD COLOR']}-{row['TAGLIA']}",
        axis=1
    )
    dati['DESCRIZIONE'] = dati_imp['Description']
    dati['STAG.'] = dati_imp['Season']
    
    # Add gender and silhouette mappings
    dati['GENERE'] = dati_imp['Gender']
    dati['TIPO'] = dati_imp['Silhouette']
    
    # Map gender to Bulgarian
    dati['GENERE PL - BG'] = dati['GENERE'].apply(map_gender_bg)
    
    # Map silhouette to Bulgarian
    dati['TIPO PER DESCRIZIONE'] = dati['TIPO'].apply(map_silhouette_bg)
    
    # Example season translations; adapt as needed
    season_translations = {
        211: 'Q1-21', 212: 'Q2-21', 213: 'Q3-21', 214: 'Q4-21',
        221: 'Q1-22', 222: 'Q2-22', 223: 'Q3-22', 224: 'Q4-22',
        231: 'Q1-23', 232: 'Q2-23', 233: 'Q3-23', 234: 'Q4-23',
        241: 'Q1-24', 242: 'Q2-24', 243: 'Q3-24', 244: 'Q4-24',
        251: 'Q1-25', 252: 'Q2-25', 253: 'Q3-25', 254: 'Q4-25',
        261: 'Q1-26', 262: 'Q2-26', 263: 'Q3-26', 264: 'Q4-26',
        271: 'Q1-27', 272: 'Q2-27', 273: 'Q3-27', 274: 'Q4-27',
        281: 'Q1-28', 282: 'Q2-28', 283: 'Q3-28', 284: 'Q4-28',
        291: 'Q1-29', 292: 'Q2-29', 293: 'Q3-29', 294: 'Q4-29',
        301: 'Q1-30', 302: 'Q2-30', 303: 'Q3-30', 304: 'Q4-30'
    }
    dati['STAG.BG'] = dati['STAG.'].map(lambda x: season_translations.get(x) if x in season_translations else str(x))

    dati['BARCODE'] = dati_imp['Barcode']
    dati['QTA'] = dati_imp['Dlv.qty']
    dati['FPC Price w/o VAT in BGN'] = dati_imp['FPC Price w/o VAT in BGN']
    dati['FPC Price +VAT in BGN'] = dati['FPC Price w/o VAT in BGN'] * 1.2
    dati['PRZ DETT'] = dati['FPC Price w/o VAT in BGN'] * 1.79
    dati['PRZ NEGOZIO'] = dati['PRZ DETT'].apply(
        lambda x: math.ceil(x / 10) * 10 - 1 if pd.notnull(x) else x
    )
    
    # Division => категория => GRUPПО mapping
    dati['категория'] = dati_imp['Division'].map({
        'EQU': 'Аксесоари',
        'APP': 'Облекло',
        'FTW': 'Обувки'
    })
    dati['ГРУПА 2'] = dati['категория'].map({
        'Обувки': 'NIKE ОБУВКИ',
        'Облекло': 'NIKE ДРЕХИ',
        'Аксесоари': 'NIKE АКСЕСОАРИ'
    })
    
    # Construct Bulgarian description for product
    dati['DESCRIZIONE BG'] = dati.apply(
        lambda row: f"{row['GENERE PL - BG']} {row['TIPO PER DESCRIZIONE']}", 
        axis=1
    )
    
    dati['Склад'] = 'BALLISTIC '
    dati['BRAND'] = 'NIKE'
    
    # Create описание сайт (product description for site/Gensoft)
    dati['Описание сайт'] = dati.apply(
        lambda row: f"{row['GENERE PL - BG']} {row['TIPO PER DESCRIZIONE']} NIKE {row['DESCRIZIONE']}", 
        axis=1
    )
    
    # 4) Build the final Gensoft dataframe
    gensoft = pd.DataFrame()
    # Non-numbered columns
    gensoft['Склад'] = dati['Склад']
    gensoft['Главна група'] = dati['BRAND']
    gensoft['Група'] = dati['ГРУПА 2']  # Use the correct 'ГРУПА 2' field
    gensoft['Стока'] = dati['COD NIKE']
    gensoft['Сер./парт. номер'] = dati['BARCODE']  # Changed from Batch to Box BarCode
    gensoft['Код на стока'] = dati['Описание сайт']
    gensoft['Баркод на стока'] = ' '  # Empty as requested
    gensoft['Мярка'] = 'бр.'
    gensoft['Количество'] = dati['QTA']
    gensoft['Доставна цена'] = dati['FPC Price +VAT in BGN']
    gensoft['Доставна валута'] = 'bgn'
    gensoft['Цена на дребно'] = dati['PRZ NEGOZIO']
    gensoft['Валута на дребно'] = 'bgn'
    gensoft['Доставчик'] = 'SPORTTIME'
    gensoft['К-во за поръчване'] = gensoft['Количество']
    gensoft['Цена'] = gensoft['Доставна цена']
    gensoft['Валута'] = 'bgn'
    gensoft['Бележка'] = dati['CODICE NIKE+COLOR']
    gensoft['Активна'] = 'Y '
    gensoft['Активна за Web'] = 'Y '
    gensoft['Ограничения в сметки'] = 'без ограничения '
    gensoft['Процент ДДС'] = '0 '
    
    # Numbered columns (previously "1 (Размер)", "2 (Пол)", etc.)
    # We'll store data exactly as before, just rename the columns to digits:
    today_str = datetime.datetime.today().strftime("%d%m%y")
    
    gensoft['1'] = ''                         # was "1 (Размер)"
    gensoft['3'] = ''                         # was "3 (Номер)"
    gensoft['14'] = dati['TAGLIA']            # was "14 (Размер сайт)"
    gensoft['107'] = dati['COD COLOR']        # was "107 (Цвят сайт)"
    gensoft['13'] = dati['CODICE BALLISTIC']  # was "13 (SKU)"
    
    # Fixed categories according to specification
    # Map Gender to Category1 correctly (column 109)
    gensoft['109'] = dati['GENERE'].apply(
        lambda x: 'Мъже' if x == 'MENS' else
                'Жени' if x == 'WOMENS' else
                'Деца' if x in ['BOYS', 'GIRLS', 'GRD SCHOOL UNS', 'BOYS GRADE SCHL'] else
                'Унисекс'  # Default for ADULT UNISEX and other cases
    )
    # Construct category 2 from gender + general category
    gensoft['110'] = dati.apply(
        lambda row: f"{row['GENERE PL - BG']} {row['категория']}", 
        axis=1
    )
    # Construct category 3 from gender + product type
    gensoft['111'] = dati.apply(
        lambda row: f"{row['GENERE PL - BG']} {row['TIPO PER DESCRIZIONE']}", 
        axis=1
    )
    
        # Map values for special columns dynamically
    gensoft['15'] = dati['BRAND']  # Brand

    # Map Gender to Column 2 (Пол) correctly
    gensoft['2'] = dati['GENERE'].apply(
        lambda x: 'Мъже' if x == 'MENS' else
                'Жени' if x == 'WOMENS' else
                'Деца' if x in ['BOYS', 'GIRLS', 'GRD SCHOOL UNS', 'BOYS GRADE SCHL'] else
                'Унисекс'  # Default for ADULT UNISEX and other cases
    )

    # Category
    gensoft['5'] = dati['категория']

    # Season
    gensoft['6'] = dati['STAG.BG']

    # Price
    gensoft['108'] = dati['PRZ NEGOZIO']

    # Description
    gensoft['106'] = dati['Описание сайт']

    # Size table code based on gender
    gensoft['113'] = dati['GENERE'].apply(
        lambda x: 'MANIKE' if x == 'MENS' else
                'WFNIKE' if x == 'WOMENS' else
                'USNIKE'  # Default for ADULT UNISEX and kids
    )

    # Import date (today)
    gensoft['112'] = datetime.datetime.today().strftime("%d%m%y")

    # Material content
    gensoft['116'] = dati_imp['Material Content']

    # Collection (could be mapped if needed)
    gensoft['7'] = 'LIFESTYLE'

    # Supplier
    gensoft['103'] = 'SPORTTIME'

    # Tags
    gensoft['120'] = ' '
    gensoft['121'] = ' '
    gensoft['122'] = ' '
    
    # 5) Define the EXACT final column order
    final_columns = [
        "Склад", "Главна група", "Група", "Стока", "Сер./парт. номер",
        "Код на стока", "Баркод на стока", "Мярка", "Количество",
        "Доставна цена", "Доставна валута", "Цена на дребно",
        "Валута на дребно", "Доставчик", "К-во за поръчване", "Цена",
        "Валута", "Бележка", "Активна", "Активна за Web",
        "Ограничения в сметки", "Процент ДДС",
        # Numbered columns in the exact order you requested:
        "1", "3", "14", "107", "13", "109", "110", "111", "15", "2", "5", "6",
        "108", "106", "113", "112", "116", "7", "103", "120", "121", "122"
    ]
    gensoft = gensoft[final_columns]
    
    # 6) Create a map for the second row
    # Only these numeric columns get a label in row 2
    two_row_header_map = {
        "1": "Размер",
        "3": "Номер",
        "14": "Размер сайт",
        "107": "Цвят сайт",
        "13": "SKU",
        "109": "Категория 1",
        "110": "Категория 2",
        "111": "Категория 3",
        "15": "Бранд",
        "2": "Пол",
        "5": "Категория",
        "6": "Сезон",
        "108": "Цена срв. сайт",
        "106": "Описание Сайт",
        "113": "Код таблица за размери",
        "112": "Дата импорт",
        "116": "Състав",
        "7": "Колекция",
        "103": "Доствчик",
        "120": "Таг1",
        "121": "Таг2",
        "122": "Таг3"
    }
    
    # 7) Write to Excel with two-row headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Gensoft"
    
    # Row 1 & 2
    for col_idx, col_name in enumerate(final_columns, start=1):
        # First row => full column name
        ws.cell(row=1, column=col_idx, value=col_name)
        
        # Second row => label only if in two_row_header_map
        if col_name in two_row_header_map:
            ws.cell(row=2, column=col_idx, value=two_row_header_map[col_name])
        else:
            ws.cell(row=2, column=col_idx, value=None)
    
    # Data rows from row 3 onward
    data_rows = dataframe_to_rows(gensoft, index=False, header=False)
    for row_idx, row_data in enumerate(data_rows, start=3):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # 8) Save the file
    wb.save(output_file)
    print(f"Transformation completed successfully.\nOutput saved to {output_file}")
    
    return gensoft